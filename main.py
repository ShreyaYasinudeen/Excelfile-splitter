import tkinter as tk
from tkinter import filedialog
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import re

def split_account(row):
    account = row['Account']
    if isinstance(account, str) and len(account) < 100:
        pattern = r'(?P<Account>\d+)(?P<Username>\D+)'
        account_data = pd.Series(account).str.extract(pattern)
        row['Account'] = account_data['Account'][0]
        row['Username'] = account_data['Username'][0]
        username = row['Username']
        print(username)
    elif isinstance(account, str) and len(account) > 100:
        trimmed_string = account[account.find("(") + 1: -1]
        parts = re.split(r",(?![^(]*\))", trimmed_string)
        usernames = []
        accounts = []
        for part in parts:
            match = re.search(r"\((.*?)\)", part)
            if match:
                account = match.group(1).strip()
                username = part[:match.start()].strip()
                usernames.append(username)
                accounts.append(account)
        return accounts, usernames
    return None, None


def process_input_file(input_file, output_folder):
    df = pd.read_excel(input_file)
    df['Username'] = split_account()
    df['FolderPath'] = df['FolderPath'].str.split('\n')
    df['FolderPath'] = df['FolderPath'].apply(lambda x: x[0] if isinstance(x, list) else x)

    modified_rows = []
    for _, row in df.iterrows():
        row_dict = row.to_dict()  # Convert row to dictionary
        accounts, usernames = split_account(row_dict)
        if accounts is not None and usernames is not None:
            for acc, user in zip(accounts, usernames):
                modified_rows.append([
                    row_dict['FolderPath'],
                    acc,
                    user,
                    row_dict['Type'],
                    row_dict['Rights']
                ])
        else:
            modified_rows.append([
                row_dict['FolderPath'],
                row_dict['Account'],
                "",
                row_dict['Type'],
                row_dict['Rights']
            ])

    modified_df = pd.DataFrame(modified_rows, columns=['FolderPath', 'Account', 'Username', 'Type', 'Rights'])


    required_columns = ['FolderPath', 'Account', 'Username', 'Type', 'Rights']

    output_file_name = os.path.splitext(os.path.basename(input_file))[0] + '_modified.xlsx'
    output_file_path = os.path.join(output_folder, output_file_name)
    modified_df[required_columns].to_excel(output_file_path, index=False)

    workbook = load_workbook(output_file_path)
    worksheet = workbook.active

    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                   min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = border

    for index, row in modified_df.iterrows():
        username = row['Username']
        if isinstance(username, list):
            username = ', '.join(username)
        worksheet.cell(row=index + 2, column=3).value = username

    workbook.save(output_file_path)
    workbook.close()

    print(f"Modified file saved to: {output_file_path}")


def upload_file():
    file_paths = filedialog.askopenfilenames(filetypes=[("Excel Files", "*.xls *.xlsx")])
    if file_paths:
        for input_file in file_paths:
            output_folder = filedialog.askdirectory()
            if output_folder:
                process_input_file(input_file, output_folder)


def submit_button_clicked():
    print("Submit button clicked!")
    upload_file()


def create_gui():
    root = tk.Tk()
    root.title("Upload Excel File")

    label = tk.Label(root, text="Select an Excel file to upload:")
    label.pack(pady=10)

    upload_button = tk.Button(root, text="Upload File", command=upload_file)
    upload_button.pack(pady=5)

    submit_button = tk.Button(root, text="Submit", command=submit_button_clicked)
    submit_button.pack(pady=5)

    root.mainloop()


if __name__ == '__main__':
    create_gui()
